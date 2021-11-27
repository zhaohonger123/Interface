# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     operate_excel
   Description :    读写Excel
   Author :       xiaoyu
   date：          2021/11/9
-------------------------------------------------
"""
import openpyxl
import os

excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "testdata", "case.xlsx")


class OperateExcel:
    def __init__(self, sheet_name, file_name=excel_path):
        """
        初始化数据
        """
        self.file = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        """
        将所有用例按字典格式转化后加入到列表data_all
        :return: data_all
        """
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet_name]
        # 通过遍历行读取Excel数据并将每行加入到一个列表row_list，ws.rows相当于一个生成器
        data_list = []
        for row in ws.rows:
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            data_list.append(row_list)
        # 将第一行与非第一行进行字典转化,并加入到列表data_all
        data_all = []
        for i in data_list[1:]:
            x = dict(zip(data_list[0], i))
            data_all.append(x)
        wb.close()
        return data_all

    def write_excel(self, row, actually, result):
        """
        将actually和result分别写入H和I列的第row行
        :param row:要写入的行
        :param result:要传入的测试结果，如:pass，failed，blocked
        :param actually:要写入的数据
        :return:
        """
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet_name]
        # self.ws.cell(row=row, column=8).value = actually
        # self.ws.cell(row=row, column=9).value = result
        ws["H{}".format(row)] = actually
        ws["I{}".format(row)] = result
        wb.save(self.file)
        wb.close()

    def write_header(self, header):
        """
        将header写入excel
        :param header: header值
        :return:
        """
        wb = openpyxl.load_workbook(self.file)
        ws = wb[self.sheet_name]
        for i in range(2, ws.max_row + 1):
            ws["E{}".format(i)] = header
        wb.save(self.file)
        wb.close()
